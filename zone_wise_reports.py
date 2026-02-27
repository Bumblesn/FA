import requests
import pandas as pd
import os
import json
import time
import logging
import datetime
from send_emails import send_email, send_failure_email
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import shutil
import pytz

# Verify datetime module
assert hasattr(datetime, 'datetime'), "datetime module is not the standard library module; check for local datetime.py"

# Set up logging
log_dir = "logs"
os.makedirs(log_dir, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(os.path.join(log_dir, f"{datetime.datetime.now().strftime('%Y%m%d')}.log")),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger()

# Initialize error log for consolidated failure email
error_log = []

# Load configuration
try:
    with open("config.json", "r") as f:
        config = json.load(f)
    api_max_retries = config.get("api_max_retries", 5)
    api_retry_delay = config.get("api_retry_delay", 2)
    base_url = config.get("base_url", "https://workozy.retool.com/api/public/4209b62f-fb47-46a5-af8f-f46dc07036a2/query?queryName=")
    admin_email = config.get("admin_email", "")
except Exception as e:
    error_log.append({
        "context": "Config Loading",
        "message": f"Error loading config.json: {str(e)}",
        "status_code": None
    })
    api_max_retries = 5
    api_retry_delay = 2
    base_url = "https://workozy.retool.com/api/public/4209b62f-fb47-46a5-af8f-f46dc07036a2/query?queryName="
    admin_email = ""

# Handle admin_email for summary report
if isinstance(admin_email, str) and admin_email:
    admin_emails = [admin_email]
elif isinstance(admin_email, list):
    admin_emails = [email for email in admin_email if email]
else:
    admin_emails = []
    logger.warning("Invalid or empty admin_email in config.json.")

# Hardcode failure email recipient
FAILURE_EMAIL = "vivek@flick2know.com"

# Determine the report date
#today = datetime.datetime.now()
LOCAL_TZ = pytz.timezone("Asia/Kolkata")
today = datetime.datetime.now(LOCAL_TZ)
if today.weekday() == 0:  # Monday
    report_date = today - datetime.timedelta(days=2)  # Saturday
else:
    report_date = today - datetime.timedelta(days=1)  # Previous day

report_date_str = report_date.strftime("%Y%m%d")
month_year = report_date.strftime("%B_%Y")

# Directory to save reports
BASE_REPORTS_DIR = "Zone_Reports"
MONTH_REPORTS_DIR = os.path.join(BASE_REPORTS_DIR, month_year)
ARCHIVE_DIR = os.path.join(BASE_REPORTS_DIR, "Archive")

# Create directories
os.makedirs(MONTH_REPORTS_DIR, exist_ok=True)
os.makedirs(ARCHIVE_DIR, exist_ok=True)

# Archive old reports (older than 3 months)
def archive_old_reports():
   #### three_months_ago = (datetime.datetime.now() - datetime.timedelta(days=90)).strftime("%B_%Y")
    three_months_ago = (datetime.datetime.now(LOCAL_TZ) - datetime.timedelta(days=90)).strftime("%B_%Y")
    for folder in os.listdir(BASE_REPORTS_DIR):
        folder_path = os.path.join(BASE_REPORTS_DIR, folder)
        if folder != "Archive" and os.path.isdir(folder_path):
            try:
                folder_date = datetime.datetime.strptime(folder, "%B_%Y")
                if folder_date < datetime.datetime.strptime(three_months_ago, "%B_%Y"):
                    archive_folder = os.path.join(ARCHIVE_DIR, folder)
                    shutil.move(folder_path, archive_folder)
                    logger.info(f"Archived {folder} to {archive_folder}")
                    print(f"Archived folder: {folder}")
            except ValueError:
                continue

print("Archiving old reports...")
archive_old_reports()
print("Archiving complete.")

# API URLs
API_ENDPOINTS = {
    "Report": "Report",
    "Zone": "Zone",
    "Employee": "Employee",
    "Beat": "Beat",
    "PrimaryCategory": "PrimaryCategory"
}

# Headers
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:137.0) Gecko/20100101 Firefox/20100101',
    'Accept': '*/*',
    'Accept-Language': 'en-US,en;q=0.5',
    'Accept-Encoding': 'gzip, deflate, br, zstd',
    'Referer': 'https://workozy.retool.com/',
    'X-Xsrf-Token': 'undefined',
    'Content-Type': 'application/json',
    'X-Retool-Client-Version': '3.186.0-7667538 (Build 240269)',
    'Origin': 'https://workozy.retool.com',
    'Connection': 'keep-alive',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-origin',
    'Priority': 'u=4',
    'TE': 'trailers'
}

# Base Payload
BASE_PAYLOAD = {
    "userParams": {
        "queryParams": {"length": 0},
        "databaseNameOverrideParams": {"length": 0},
        "databaseHostOverrideParams": {"length": 0},
        "databaseUsernameOverrideParams": {"length": 0},
        "databasePasswordOverrideParams": {"length": 0}
    },
    "password": "",
    "environment": "production",
    "queryType": "SqlQueryUnified",
    "frontendVersion": "1",
    "releaseVersion": None,
    "includeQueryExecutionMetadata": True,
    "streamResponse": False
}

# Load email configuration
def load_email_config():
    try:
        with open("email_config.json", "r") as f:
            return json.load(f)
    except Exception as e:
        error_log.append({
            "context": "Email Config Loading",
            "message": f"Error loading email_config.json: {str(e)}",
            "status_code": None
        })
        print(f"Error loading email_config.json: {str(e)}")
        return {}

# Format Excel file
def format_excel(filename):
    try:
        wb = load_workbook(filename)
        ws = wb.active
        # Freeze top row
        ws.freeze_panes = "A2"
        # Format header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        # Format numeric columns as integers
        for col in ws.iter_cols(min_row=2):
            header = ws.cell(1, col[0].column).value
            if header and ws.cell(2, col[0].column).value is not None:
                try:
                    float(ws.cell(2, col[0].column).value)  # Check if numeric
                    for cell in col:
                        cell.number_format = "0"
                except (ValueError, TypeError):
                    pass
        wb.save(filename)
        logger.info(f"Formatted Excel file: {filename}")
        return True
    except Exception as e:
        error_log.append({
            "context": f"Excel Formatting ({filename})",
            "message": f"Error formatting {filename}: {str(e)}",
            "status_code": None
        })
        print(f"Error formatting {filename}: {str(e)}")
        return False

# Function to fetch data from APIs
def fetch_data(api_name, zone_id=None):
    url = base_url + api_name
    payload = BASE_PAYLOAD.copy()
    
    if api_name == "Report" and zone_id is not None:
        payload["userParams"]["queryParams"] = {
            "0": report_date_str,
            "1": report_date_str,
            "2": str(zone_id),
            "3": report_date_str,
            "4": report_date_str,
            "5": str(zone_id),
            "6": report_date_str,
            "7": report_date_str,
            "8": str(zone_id),
            "length": 9
        }
    
    for attempt in range(1, api_max_retries + 1):
        try:
            response = requests.post(url, headers=HEADERS, json=payload)
            logger.info(f"Fetching {api_name} API{' for ZoneId ' + str(zone_id) if zone_id else ''} - Attempt {attempt}/{api_max_retries} - Status Code: {response.status_code}")
            
            if response.status_code == 200:
                try:
                    data = response.json()
                except requests.exceptions.JSONDecodeError as e:
                    error_log.append({
                        "context": f"Fetch {api_name} (ZoneId: {zone_id})",
                        "message": f"Failed to parse JSON response: {str(e)}",
                        "status_code": response.status_code
                    })
                    return pd.DataFrame()
                
                if api_name == "Zone" and "queryData" in data and "zId" in data["queryData"]:
                    zone_data = [{"zId": zid, "zName": zname} for zid, zname in zip(data["queryData"]["zId"], data["queryData"]["zName"])]
                    return pd.DataFrame(zone_data)
                if api_name == "Report" and "queryData" in data:
                    report_data = [
                        {key: values[i] for key, values in data["queryData"].items()}
                        for i in range(len(next(iter(data["queryData"].values()))))
                    ]
                    df = pd.DataFrame(report_data)
                    logger.info(f"Fetched {api_name} for ZoneId {zone_id}: {df.shape[0]} rows")
                    return df
                return pd.DataFrame(data.get("queryData", []))
            
            error_log.append({
                "context": f"Fetch {api_name} (ZoneId: {zone_id})",
                "message": f"HTTP error: {response.text}",
                "status_code": response.status_code
            })
        
        except Exception as e:
            error_log.append({
                "context": f"Fetch {api_name} (ZoneId: {zone_id})",
                "message": f"Exception on attempt {attempt}/{api_max_retries}: {str(e)}",
                "status_code": None
            })
        
        if attempt < api_max_retries:
            logger.info(f"Retrying {api_name} in {api_retry_delay} seconds...")
            time.sleep(api_retry_delay)
    
    error_log.append({
        "context": f"Fetch {api_name} (ZoneId: {zone_id})",
        "message": f"Failed to fetch after {api_max_retries} attempts",
        "status_code": None
    })
    return pd.DataFrame()

# Fetch non-report data
print("Fetching Zone data...")
df_zone = fetch_data("Zone")
print("Zone data fetch complete.")

df_employee = fetch_data("Employee")
print("Employee data fetched.")

df_beat = fetch_data("Beat")
print("Beat data fetched.")

df_primary_category = fetch_data("PrimaryCategory")
print("PrimaryCategory data fetched.")

# Ensure Zone API has data
if df_zone.empty:
    error_log.append({
        "context": "Zone Data Check",
        "message": "Zone data is empty after retries.",
        "status_code": None
    })
    logger.error("Zone data is empty after retries. Exiting without generating or emailing reports.")
    print("ERROR: Zone data is empty. Exiting.")
    if FAILURE_EMAIL:
        try:
            send_failure_email("Zone API Failure", "Failed to fetch Zone data after retries.", [FAILURE_EMAIL])
            logger.info("Sent Zone API failure notification to admin.")
            print("Sent Zone API failure email to admin.")
        except Exception as e:
            error_log.append({
                "context": "Zone API Failure Email",
                "message": f"Failed to send Zone API failure email: {str(e)}",
                "status_code": None
            })
    exit()

# 🔹 Preprocess Zone Data
df_zone.rename(columns={"zId": "ZoneId", "zName": "Zone"}, inplace=True)

# 🔹 Get unique zones
unique_zones = df_zone[['ZoneId', 'Zone']].drop_duplicates()

# 🔹 Preprocess Employee Data
if not df_employee.empty:
    df_employee["ESMId"] = df_employee["Id"].apply(lambda x: x[0] if isinstance(x, list) and x else x)
    df_employee["User"] = df_employee["Cname"].apply(lambda x: x[0] if isinstance(x, list) and x else x)

# Load email configuration
print("Loading email configuration...")
email_config = load_email_config()
print("Email configuration loaded.")

# Collect all zone reports for summary
all_zone_reports = []

# 🔹 Process reports for each zone
for _, zone_row in unique_zones.iterrows():
    zone_id = zone_row['ZoneId']
    zone_name = zone_row['Zone']
    
    print(f"Starting processing for Zone: {zone_name} (ZoneId: {zone_id})")
    
    try:
        # Fetch report data
        print(f"Fetching data for Zone: {zone_name}")
        df_zone_report = fetch_data("Report", zone_id=zone_id)
        
        if df_zone_report.empty:
            logger.warning(f"No data found for Zone: {zone_name} after retries. Skipping...")
            print(f"No data found for Zone: {zone_name}. Skipping.")
            error_log.append({
                "context": f"Zone Report ({zone_name})",
                "message": f"No data found for ZoneId {zone_id} after retries.",
                "status_code": None
            })
            continue
        
        print(f"Data fetched successfully for Zone: {zone_name} ({df_zone_report.shape[0]} rows)")
        
        # 🔹 Preprocess Report Data
        if "ESMId" in df_zone_report.columns:
            df_zone_report["ESMId"] = df_zone_report["ESMId"].apply(lambda x: x[0] if isinstance(x, list) and x else x)
        
        # 🔹 Data Processing
        df_zone_report = df_zone_report.merge(df_zone[['ZoneId', 'Zone']], on="ZoneId", how="left").drop(columns=["ZoneId"], errors="ignore")
        
        if not df_employee.empty and "ESMId" in df_zone_report.columns:
            df_zone_report = df_zone_report.merge(df_employee[['ESMId', 'User']], on="ESMId", how="left").drop(columns=["ESMId"], errors="ignore")
        
        if not df_beat.empty and "SelectedBeatId" in df_zone_report.columns:
            df_beat.rename(columns={"Id": "SelectedBeatId", "BeatName": "Selected Beat"}, inplace=True)
            df_zone_report = df_zone_report.merge(df_beat[['SelectedBeatId', 'Selected Beat']], on="SelectedBeatId", how="left").drop(columns=["SelectedBeatId"], errors="ignore")
        
        if "JW" in df_zone_report.columns:
            df_zone_report['Joint Working'] = df_zone_report['JW'].apply(lambda x: 'Yes' if x == "Yes" else 'No')
            df_zone_report.drop(columns=["JW"], inplace=True, errors="ignore")
        
        if "JointWorkingEmployeeId" in df_zone_report.columns and not df_employee.empty:
            df_zone_report = df_zone_report.merge(df_employee[['ESMId', 'User']],
                                                 left_on="JointWorkingEmployeeId",
                                                 right_on="ESMId",
                                                 how="left")
            if "User_x" in df_zone_report.columns:
                df_zone_report.rename(columns={"User_x": "User"}, inplace=True)
            if "User_y" in df_zone_report.columns:
                df_zone_report.rename(columns={"User_y": "JointWorkingEmployeeName"}, inplace=True)
            df_zone_report.drop(columns=["JointWorkingEmployeeId", "ESMId"], inplace=True, errors="ignore")
        
        # 🔹 Rename columns
        df_zone_report.rename(columns={
            "DayStartDateKey": "Date",
            "ESMRank": "User Rank",
            "JointWorkingCalls": "JW TC",
            "TC": "TC",
            "PCInJointWorking": "JW PC",
            "PC": "PC",
            "LEDBULBQty": "LED BULB Qty",
            "JWLEDBULBQty": "JW LED BULB Qty",
            "LEDBattenQty": "LED Batten Qty",
            "JWLEDBattenQty": "JW LED Batten Qty",
            "LEDDOWNLIGHTNetValue": "LED DOWNLIGHT TotalValue",
            "JWLEDDOWNLIGHTNetValue": "JW LED DOWNLIGHT TotalValue",
            "MCBQty": "MCB Qty",
            "JWMCBQty": "JW MCB Qty",
            "TotalValue": "Total Value",
            "JWTotalValue": "JW Total Value",
            "TotalNetValue": "Total Net Value",
            "JWTotalNetValue": "JW Total Net Value"
        }, inplace=True)
        
        # 🔹 Update User Rank
        if "User Rank" in df_zone_report.columns:
            df_zone_report["User Rank"] = df_zone_report["User Rank"].replace({"ESM": "Employee", "ASM": "SO"})
        
        # 🔹 Define final headers
        final_headers = [
            "Date", "Zone", "User", "User Rank", "JointWorkingEmployeeName", "Joint Working",
            "Selected Beat", "JW TC", "TC", "JW PC", "PC", "LED BULB Qty", "JW LED BULB Qty",
            "LED Batten Qty", "JW LED Batten Qty", "LED DOWNLIGHT TotalValue", "JW LED DOWNLIGHT TotalValue",
            "MCB Qty", "JW MCB Qty", "Total Net Value", "JW Total Net Value"
        ]
        
        # Ensure all columns exist
        for column in final_headers:
            if column not in df_zone_report.columns:
                df_zone_report[column] = ""
        
        df_zone_report = df_zone_report[final_headers]
        
        # 🔹 Convert numeric columns to integers
        numeric_columns = ["JW TC", "TC", "JW PC", "PC", "LED BULB Qty", "JW LED BULB Qty",
                           "LED Batten Qty", "JW LED Batten Qty", "LED DOWNLIGHT TotalValue", "JW LED DOWNLIGHT TotalValue",
                           "MCB Qty", "JW MCB Qty", "Total Net Value", "JW Total Net Value"]
        for col in numeric_columns:
            if col in df_zone_report.columns:
                df_zone_report[col] = pd.to_numeric(df_zone_report[col], errors='coerce').round().fillna(0).astype(int)
        
        logger.info(f"Processed report for {zone_name}: {df_zone_report.shape[0]} rows")
        
        # 🔹 Save Zone Report
        safe_zone_name = "".join(c for c in zone_name if c.isalnum() or c in (' ', '_')).strip().replace(" ", "_")
        filename = os.path.join(MONTH_REPORTS_DIR, f"Manager_Working_report_{safe_zone_name}_{report_date_str}.xlsx")
        
        print(f"Saving report for Zone: {zone_name}")
        try:
            df_zone_report.to_excel(filename, index=False)
            if format_excel(filename):
                print(f"Report saved for Zone: {zone_name} at {filename}")
                logger.info(f"Report saved as {filename}")
            else:
                error_log.append({
                    "context": f"Zone Report Save ({zone_name})",
                    "message": f"Failed to format Excel for {zone_name}",
                    "status_code": None
                })
                continue
        except Exception as e:
            error_log.append({
                "context": f"Zone Report Save ({zone_name})",
                "message": f"Error saving Excel for {zone_name}: {str(e)}",
                "status_code": None
            })
            print(f"Failed to save report for Zone: {zone_name}: {str(e)}")
            continue
        
        # 🔹 Collect for summary
        all_zone_reports.append(df_zone_report)
        
        # 🔹 Send Email
        zone_email_config = email_config.get(safe_zone_name, {})
        to_recipients = [email for email in zone_email_config.get("to", []) if email]
        cc_recipients = [email for email in zone_email_config.get("cc", []) if email]
        
        if to_recipients:
            print(f"Sending email for Zone: {zone_name}")
            try:
                send_email(zone_name, filename, to_recipients, cc_recipients)
                print(f"Email sent for Zone: {zone_name}")
            except Exception as e:
                error_log.append({
                    "context": f"Zone Email ({zone_name})",
                    "message": f"Failed to send email for {zone_name}: {str(e)}",
                    "status_code": None
                })
                print(f"Failed to send email for Zone: {zone_name}: {str(e)}")
        else:
            logger.warning(f"No valid email recipients for Zone: {zone_name}. Skipping email.")
            print(f"No valid email recipients for Zone: {zone_name}. Skipping email.")
            error_log.append({
                "context": f"Zone Email ({zone_name})",
                "message": f"No valid email recipients for Zone: {zone_name}.",
                "status_code": None
            })
        
        print(f"Completed processing for Zone: {zone_name}")
    
    except Exception as e:
        error_log.append({
            "context": f"Zone Processing ({zone_name})",
            "message": f"Error processing Zone: {zone_name}: {str(e)}",
            "status_code": None
        })
        logger.error(f"Error processing Zone: {zone_name}: {str(e)}")
        print(f"Failed processing for Zone: {zone_name}: {str(e)}")
        continue

# 🔹 Generate Summary Report
print("Generating summary report...")
if all_zone_reports:
    try:
        summary_df = pd.concat(all_zone_reports, ignore_index=True)
        summary_filename = os.path.join(MONTH_REPORTS_DIR, f"Summary_Working_report_{report_date_str}.xlsx")
        
        # Create pivot table
        pivot = pd.pivot_table(
            summary_df,
            values=["TC", "PC", "Total Net Value"],
            index=["Zone"],
            aggfunc="sum",
            fill_value=0
        ).reset_index()
        
        # Convert numeric columns to integers
        for col in ["TC", "PC", "Total Net Value"]:
            if col in pivot.columns:
                pivot[col] = pd.to_numeric(pivot[col], errors='coerce').round().fillna(0).astype(int)
        
        # Save summary
        print(f"Saving summary report...")
        pivot.to_excel(summary_filename, index=False)
        if format_excel(summary_filename):
            logger.info(f"Summary report saved as {summary_filename}")
            print(f"Summary report saved at {summary_filename}")
        else:
            error_log.append({
                "context": "Summary Report Save",
                "message": "Failed to format summary report Excel.",
                "status_code": None
            })
            print(f"Failed to format summary report")
        
        # Email summary to admin only
        if admin_emails:
            print("Sending summary report email...")
            try:
                send_email("All Zones Summary", summary_filename, admin_emails, [])
                logger.info("Summary report email sent.")
                print("Summary report email sent.")
            except Exception as e:
                error_log.append({
                    "context": "Summary Report Email",
                    "message": f"Failed to send summary email: {str(e)}",
                    "status_code": None
                })
                print(f"Failed to send summary email: {str(e)}")
        else:
            logger.warning("No valid admin_emails configured. Skipping summary report email.")
            print("No valid admin_emails configured. Skipping summary report email.")
            error_log.append({
                "context": "Summary Report Email",
                "message": "No valid admin_emails configured.",
                "status_code": None
            })
    except Exception as e:
        error_log.append({
            "context": "Summary Report Generation",
            "message": f"Error generating summary report: {str(e)}",
            "status_code": None
        })
        print(f"Error generating summary report: {str(e)}")
else:
    logger.warning("No zone reports generated. Skipping summary report.")
    print("No zone reports generated. Skipping summary report.")
    error_log.append({
        "context": "Summary Report Generation",
        "message": "No zone reports generated.",
        "status_code": None
    })

# 🔹 Send Consolidated Failure Email
if error_log:
    print("Sending consolidated failure email...")
    try:
        # Format error message
        error_message = "The following errors occurred during script execution:\n\n"
        for error in error_log:
            error_message += f"**Context**: {error['context']}\n"
            error_message += f"**Message**: {error['message']}\n"
            error_message += f"**Status Code**: {error['status_code'] if error['status_code'] is not None else 'N/A'}\n\n"
        
        send_failure_email("Zone Reports Script Failures", error_message, [FAILURE_EMAIL])
        logger.info(f"Consolidated failure email sent to {FAILURE_EMAIL}")
        print(f"Consolidated failure email sent to {FAILURE_EMAIL}")
    except Exception as e:
        logger.error(f"Failed to send consolidated failure email: {str(e)}")
        print(f"Failed to send consolidated failure email: {str(e)}")

logger.info("All zone reports and summary generated and emails sent successfully.")

print("Script execution completed.")
